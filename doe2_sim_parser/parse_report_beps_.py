import pprint
import re
from collections import OrderedDict, namedtuple
from itertools import starmap
from typing import Generator, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet import Worksheet

Meter = namedtuple("Meter", ["name", "type_"])
Categories = (
    "LIGHTS",
    "TASK\nLIGHTS",
    "MISC\nEQUIP",
    "SPACE\nHEATING",
    "SPACE\nCOOLING",
    "HEAT\nREJECT",
    "PUMPS\n& AUX",
    "VENT\nFANS",
    "REFRIG\nDISPLAY",
    "HT PUMP\nSUPPLEN",
    "DOMEST\nHOT WTR",
    "EXT\nUSAGE",
    "TOTAL",
)

pattern_meter = re.compile(
    r"""(?P<name>.+)\s{2}(?P<type_>ELECTRICITY|NATURAL\-GAS)""",
    flags=re.VERBOSE,
)

pattern_no_by_category = re.compile(
    r"""
    ^\s+(?P<unit>[^\s]+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    \s+(\d+\.\d+)
    """,
    flags=re.VERBOSE,
)

pattern_total_energy = re.compile(
    r"""
\s+
(?P<name>TOTAL\sSITE\sENERGY|TOTAL\sSOURCE\sENERGY)
\s+
(?P<value>\d+\.\d+)
\s
(?P<unit>MBTU)
\s+
(?P<value_per_gross_area>\d+\.\d+)
\s+
(?P<unit_per_gross_area>KBTU\/SQFT\-YR\sGROSS\-AREA)
\s+
(?P<value_per_net_area>\d+\.\d+)
\s+
(?P<unit_per_net_area>KBTU\/SQFT\-YR\sNET\-AREA)
""",
    flags=re.VERBOSE,
)

pattern_percent_and_hours = re.compile(
    r"""
\s+
(?P<name>.+?)
\s+
=
\s+
(?P<value>\d+[.\d]*)
""",
    flags=re.VERBOSE,
)
pp = pprint.PrettyPrinter(indent=4, width=180)


def chunks(list_: Iterable, n: int) -> Generator[List, None, None]:
    if list_:
        yield list_[:n]
        yield from chunks(list_[n:], n)


def parse_meter(line: str) -> Meter:
    _ = pattern_meter.search(line)
    return Meter(**_.groupdict())


def parse_no_by_category(line: str) -> OrderedDict:
    _ = pattern_no_by_category.search(line)

    return OrderedDict((
        *_.groupdict().items(),
        *starmap(lambda x, y: (x, float(y)), zip(Categories,
                                                 _.groups()[1:])),
    ))


def parse_content(lines: List[str]) -> Tuple:
    return tuple(
        map(
            lambda x: (parse_meter(x[0]), parse_no_by_category(x[1])),
            chunks(list(filter(lambda x: x.strip(), lines)), 2),
        ))


def parse_sum(lines: List[str]) -> Tuple:
    return tuple(map(lambda x: parse_no_by_category(x), lines))


def parse_total(lines: List[str]):
    return tuple(map(lambda x: pattern_total_energy.search(x).groups(), lines))


def parse_percent(lines: List[str]):
    return tuple(
        list(
            map(lambda x: pattern_percent_and_hours.search(x).groupdict(),
                lines)))


def write_categories(ws: Worksheet, cell: Cell, categories: List[str]):
    _cell: Cell = None
    for index, value in enumerate((*[None] * 3, *categories)):
        _cell = ws.cell(row=cell.row, column=cell.col_idx + index)
        _cell.value = value
    return ws.cell(row=_cell.row + 1, column=1)


def write_content(ws: Worksheet, cell: Cell, content: List[str]):
    _cell: Cell = ws.cell(row=cell.row, column=1)
    for i, v in enumerate(content):
        meter, value = v

        ws.cell(row=_cell.row, column=_cell.col_idx).value = meter.name
        ws.cell(row=_cell.row, column=_cell.col_idx + 1).value = meter.type_

        for index, _value in enumerate(value.items()):
            ws.cell(
                row=_cell.row,
                column=_cell.col_idx + 2 + index).value = _value[1]

        _cell = ws.cell(row=_cell.row + 1, column=1)
    return _cell


def write_sum(ws: Worksheet, cell: Cell, sum_: List[str]):
    _cell: Cell = ws.cell(row=cell.row, column=1)
    for _ in sum_:
        for i, pair in enumerate(_.items()):
            k, v = pair
            ws.cell(row=_cell.row, column=_cell.col_idx + 2 + i).value = v
        _cell = ws.cell(row=_cell.row + 1, column=1)
    return _cell


def write_total(ws: Worksheet, cell: Cell, total: List[str]):
    _cell: Cell = ws.cell(row=cell.row, column=1)
    for _ in total:
        for i, v in enumerate(_):
            ws.cell(row=_cell.row, column=_cell.col_idx + i).value = v
        _cell = ws.cell(row=_cell.row + 1, column=1)
    return _cell


def write_percent(ws: Worksheet, cell: Cell, percent: List[str]):
    _cell: Cell = ws.cell(row=cell.row, column=1)
    for _ in percent:
        ws.cell(row=_cell.row, column=_cell.col_idx).value = _["name"]
        ws.cell(row=_cell.row, column=_cell.col_idx + 1).value = _["value"]
        _cell = ws.cell(row=_cell.row + 1, column=1)
    return _cell


def write_note(ws: Worksheet, cell: Cell, note: List[str]):
    _cell: Cell = ws.cell(row=cell.row, column=1)
    for _ in note:
        ws.cell(row=_cell.row, column=_cell.col_idx).value = _.strip()
        _cell = ws.cell(row=_cell.row + 1, column=1)
    return _cell


SliceFunc = namedtuple("SliceFunc",
                       ["name", "slice", "func_parse", "func_write"])

Write = namedtuple("Write", ["result", "func_write"])

SLICES_BEPS = (
    SliceFunc(
        name="header",
        slice=slice(0, 3),
        func_parse=lambda x: x,
        func_write=None),
    SliceFunc(
        name="categories",
        slice=slice(5, 7),
        func_parse=lambda x: Categories,
        func_write=write_categories,
    ),
    SliceFunc(
        name="content",
        slice=slice(9, -17),
        func_parse=parse_content,
        func_write=write_content,
    ),
    SliceFunc(
        name="summary",
        slice=slice(-15, -14),
        func_parse=parse_sum,
        func_write=write_sum,
    ),
    SliceFunc(
        name="total",
        slice=slice(-11, -9),
        func_parse=parse_total,
        func_write=write_total,
    ),
    SliceFunc(
        name="percent",
        slice=slice(-8, -4),
        func_parse=parse_percent,
        func_write=write_percent,
    ),
    SliceFunc(
        name="note",
        slice=slice(-3, -2),
        func_parse=lambda x: x,
        func_write=write_note),
)


def parse_beps(report: List[str]):
    yield from map(
        lambda x: Write(result=x.func_parse(report[x.slice]), func_write=x.func_write),
        SLICES_BEPS,
    )


def write_beps(wb: Workbook, report: List[str]):
    ws: Worksheet = wb.create_sheet("BEPS")
    cell_anchor: Cell = ws.cell(row=1, column=1)

    for _ in parse_beps(report):
        if _.func_write:
            cell_anchor = _.func_write(ws, cell_anchor, _.result)

    return cell_anchor
